from __future__ import annotations

import tempfile
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable, List, Optional, Sequence

import cv2
import numpy as np
from openpyxl import Workbook

from engines.paddle_internal.layout_analyzer import LayoutRegion


@dataclass(frozen=True)
class TableData:
    page_number: int
    table_index: int
    rows: list[list[str]]

    @property
    def sheet_name(self) -> str:
        return f"table_{self.page_number}_{self.table_index}"


class TableExtractionError(Exception):
    """Raised when table extraction cannot proceed."""


class _SimpleHTMLTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._current_row: list[str] = []
        self._cell_chunks: list[str] = []
        self._in_cell = False

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag == "tr":
            self._current_row = []
        elif tag in {"td", "th"}:
            self._in_cell = True
            self._cell_chunks = []

    def handle_data(self, data: str) -> None:
        if self._in_cell:
            self._cell_chunks.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"td", "th"}:
            self._in_cell = False
            text = " ".join(chunk.strip() for chunk in self._cell_chunks if chunk.strip())
            self._current_row.append(text)
            self._cell_chunks = []
        elif tag == "tr":
            if self._current_row:
                self.rows.append(self._current_row)
            self._current_row = []



def extract_tables(
    regions: Sequence[LayoutRegion],
    image_paths: dict[int, Path] | None = None,
) -> list[TableData]:
    tables: list[TableData] = []
    table_counts: dict[int, int] = {}

    # OCRエンジンを1回だけ初期化して全テーブルで使い回す
    ocr_engine = None
    if image_paths:
        try:
            from paddleocr import PaddleOCR
            ocr_engine = PaddleOCR(lang="japan", use_angle_cls=True)
        except Exception:
            pass

    for region in regions:
        if region.region_type not in {"table", "table_frame_candidate"}:
            continue

        table_counts[region.page_number] = table_counts.get(region.page_number, 0) + 1
        table_index = table_counts[region.page_number]

        # まず既存の HTML / text から行を取得
        rows = _extract_rows_from_region(region.raw)

        # 空の場合、グリッド線 → セル単位 OCR で再構成
        if not rows and ocr_engine and image_paths and region.bbox:
            img_path = image_paths.get(region.page_number)
            if img_path:
                rows = _extract_rows_from_image_crop(img_path, region.bbox, ocr_engine)

        if rows:
            tables.append(TableData(page_number=region.page_number, table_index=table_index, rows=rows))

    return tables



def write_tables_to_workbook(tables: Sequence[TableData], output_path: str | Path) -> Optional[Path]:
    if not tables:
        return None

    path = Path(output_path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for table in tables:
        sheet = workbook.create_sheet(title=table.sheet_name[:31])
        for row_index, row in enumerate(table.rows, start=1):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)

    workbook.save(path)
    return path



def _extract_rows_from_region(raw_region: dict[str, Any]) -> list[list[str]]:
    html = raw_region.get("res", {}).get("html")
    if isinstance(html, str) and html.strip():
        rows = _parse_html_table(html)
        if rows:
            return rows

    text = raw_region.get("res", {}).get("text")
    if isinstance(text, list):
        rows = [[str(cell) for cell in row] for row in text if isinstance(row, list)]
        if rows:
            return rows

    return []



def _parse_html_table(html: str) -> list[list[str]]:
    parser = _SimpleHTMLTableParser()
    parser.feed(html)
    parser.close()
    return parser.rows


def _extract_rows_from_image_crop(
    image_path: Path,
    bbox: list,
    ocr_engine: Any = None,
) -> list[list[str]]:
    """表領域をグリッド線で分割してセル単位にOCRし、行列を再構成する。"""
    from engines.paddle_internal.frame_detector import detect_grid_lines

    image = cv2.imread(str(image_path))
    if image is None:
        return []

    h, w = image.shape[:2]
    x1, y1 = max(0, int(bbox[0])), max(0, int(bbox[1]))
    x2, y2 = min(w, int(bbox[2])), min(h, int(bbox[3]))
    if x2 <= x1 or y2 <= y1:
        return []

    int_bbox = [x1, y1, x2, y2]
    h_lines, v_lines = detect_grid_lines(image_path, int_bbox)

    # グリッド線が十分あればセル単位OCR、なければbboxクラスタリングにフォールバック
    if len(h_lines) >= 2 and len(v_lines) >= 2:
        return _ocr_grid_cells(image, int_bbox, h_lines, v_lines, ocr_engine)
    else:
        return _ocr_bbox_clustering(image, int_bbox, ocr_engine)


def _ocr_grid_cells(
    image: np.ndarray,
    bbox: list[int],
    h_lines: list[int],
    v_lines: list[int],
    ocr_engine: Any,
) -> list[list[str]]:
    """罫線で定義されたセルをひとつずつOCRして表を再構成する。"""
    x1, y1 = bbox[0], bbox[1]
    rows: list[list[str]] = []
    pad = 2

    for r in range(len(h_lines) - 1):
        row: list[str] = []
        cy1 = y1 + h_lines[r] + pad
        cy2 = y1 + h_lines[r + 1] - pad

        for c in range(len(v_lines) - 1):
            cx1 = x1 + v_lines[c] + pad
            cx2 = x1 + v_lines[c + 1] - pad

            cell_img = image[cy1:cy2, cx1:cx2]
            if cell_img.size == 0 or cell_img.shape[0] < 5 or cell_img.shape[1] < 5:
                row.append("")
                continue

            text = _ocr_single_cell(cell_img, ocr_engine)
            row.append(text)

        if any(cell.strip() for cell in row):
            rows.append(row)

    return rows


def _ocr_bbox_clustering(
    image: np.ndarray,
    bbox: list[int],
    ocr_engine: Any,
) -> list[list[str]]:
    """グリッド線が取れない場合、OCRのbbox y座標でクラスタリングして行を作る。"""
    x1, y1, x2, y2 = bbox
    cropped = image[y1:y2, x1:x2]

    tmp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp_path = Path(tmp.name)
        cv2.imwrite(str(tmp_path), cropped)
        result = ocr_engine.ocr(str(tmp_path), cls=True)
    except Exception:
        return []
    finally:
        if tmp_path and tmp_path.exists():
            tmp_path.unlink(missing_ok=True)

    if not result or not result[0]:
        return []

    items: list[tuple[float, float, str]] = []
    for line in result[0]:
        if not line or len(line) < 2:
            continue
        box, rec = line[0], line[1]
        text = rec[0] if isinstance(rec, (list, tuple)) else str(rec)
        if not text.strip():
            continue
        y_center = (box[0][1] + box[2][1]) / 2
        x_center = (box[0][0] + box[2][0]) / 2
        items.append((y_center, x_center, text.strip()))

    if not items:
        return []

    items.sort(key=lambda item: item[0])
    y_values = [item[0] for item in items]
    row_gap = max((max(y_values) - min(y_values)) / max(len(y_values), 1) * 1.5, 20.0)

    row_groups: list[list[tuple[float, float, str]]] = []
    current = [items[0]]
    for item in items[1:]:
        if item[0] - current[-1][0] <= row_gap:
            current.append(item)
        else:
            row_groups.append(current)
            current = [item]
    row_groups.append(current)

    return [[text for _, _, text in sorted(row, key=lambda r: r[1])] for row in row_groups]


def _ocr_single_cell(cell_img: np.ndarray, ocr_engine: Any) -> str:
    """セル画像をOCRして文字列を返す。"""
    tmp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp_path = Path(tmp.name)
        cv2.imwrite(str(tmp_path), cell_img)
        result = ocr_engine.ocr(str(tmp_path), cls=True)
    except Exception:
        return ""
    finally:
        if tmp_path and tmp_path.exists():
            tmp_path.unlink(missing_ok=True)

    if not result or not result[0]:
        return ""

    texts = []
    for line in result[0]:
        if not line or len(line) < 2:
            continue
        rec = line[1]
        text = rec[0] if isinstance(rec, (list, tuple)) else str(rec)
        if text.strip():
            texts.append(text.strip())
    return " ".join(texts)
