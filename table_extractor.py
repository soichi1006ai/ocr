from __future__ import annotations

import tempfile
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable, List, Optional, Sequence

import cv2
import numpy as np
from openpyxl import Workbook

from layout_analyzer import LayoutRegion


@dataclass(frozen=True)
class TableData:
    page_number: int
    table_index: int
    rows: list[list[str]]

    @property
    def sheet_name(self) -> str:
        return f"table_{self.page_number}_{self.table_index}"


class TableExtractionError(Exception):
    """Raised when table extraction cannot proceed."""


class _SimpleHTMLTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._current_row: list[str] = []
        self._cell_chunks: list[str] = []
        self._in_cell = False

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag == "tr":
            self._current_row = []
        elif tag in {"td", "th"}:
            self._in_cell = True
            self._cell_chunks = []

    def handle_data(self, data: str) -> None:
        if self._in_cell:
            self._cell_chunks.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"td", "th"}:
            self._in_cell = False
            text = " ".join(chunk.strip() for chunk in self._cell_chunks if chunk.strip())
            self._current_row.append(text)
            self._cell_chunks = []
        elif tag == "tr":
            if self._current_row:
                self.rows.append(self._current_row)
            self._current_row = []



def extract_tables(
    regions: Sequence[LayoutRegion],
    image_paths: dict[int, Path] | None = None,
) -> list[TableData]:
    tables: list[TableData] = []
    table_counts: dict[int, int] = {}

    for region in regions:
        if region.region_type not in {"table", "table_frame_candidate"}:
            continue

        table_counts[region.page_number] = table_counts.get(region.page_number, 0) + 1
        table_index = table_counts[region.page_number]

        # まず既存の HTML / text から行を取得
        rows = _extract_rows_from_region(region.raw)

        # 空の場合、画像クロップ + OCR で再構成
        if not rows and image_paths and region.bbox:
            img_path = image_paths.get(region.page_number)
            if img_path:
                rows = _extract_rows_from_image_crop(img_path, region.bbox)

        if rows:
            tables.append(TableData(page_number=region.page_number, table_index=table_index, rows=rows))

    return tables



def write_tables_to_workbook(tables: Sequence[TableData], output_path: str | Path) -> Optional[Path]:
    if not tables:
        return None

    path = Path(output_path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for table in tables:
        sheet = workbook.create_sheet(title=table.sheet_name[:31])
        for row_index, row in enumerate(table.rows, start=1):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)

    workbook.save(path)
    return path



def _extract_rows_from_region(raw_region: dict[str, Any]) -> list[list[str]]:
    html = raw_region.get("res", {}).get("html")
    if isinstance(html, str) and html.strip():
        rows = _parse_html_table(html)
        if rows:
            return rows

    text = raw_region.get("res", {}).get("text")
    if isinstance(text, list):
        rows = [[str(cell) for cell in row] for row in text if isinstance(row, list)]
        if rows:
            return rows

    return []



def _parse_html_table(html: str) -> list[list[str]]:
    parser = _SimpleHTMLTableParser()
    parser.feed(html)
    parser.close()
    return parser.rows


def _extract_rows_from_image_crop(image_path: Path, bbox: list) -> list[list[str]]:
    """表領域をクロップし PaddleOCR で認識、bbox の y/x 座標から行列を再構成する。"""
    image = cv2.imread(str(image_path))
    if image is None:
        return []

    h, w = image.shape[:2]
    x1 = max(0, int(bbox[0]))
    y1 = max(0, int(bbox[1]))
    x2 = min(w, int(bbox[2]))
    y2 = min(h, int(bbox[3]))
    if x2 <= x1 or y2 <= y1:
        return []

    cropped = image[y1:y2, x1:x2]

    tmp_path: Path | None = None
    try:
        from paddleocr import PaddleOCR
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp_path = Path(tmp.name)
        cv2.imwrite(str(tmp_path), cropped)

        ocr = PaddleOCR(lang="japan", use_angle_cls=True)
        result = ocr.ocr(str(tmp_path), cls=True)
    except Exception:
        return []
    finally:
        if tmp_path and tmp_path.exists():
            tmp_path.unlink(missing_ok=True)

    if not result or not result[0]:
        return []

    # (y中心, x中心, テキスト) のリストを作成
    items: list[tuple[float, float, str]] = []
    for line in result[0]:
        if not line or len(line) < 2:
            continue
        box, rec = line[0], line[1]
        text = rec[0] if isinstance(rec, (list, tuple)) else str(rec)
        if not text.strip():
            continue
        y_center = (box[0][1] + box[2][1]) / 2
        x_center = (box[0][0] + box[2][0]) / 2
        items.append((y_center, x_center, text.strip()))

    if not items:
        return []

    # y 座標でソートして行をクラスタリング
    items.sort(key=lambda item: item[0])
    y_values = [item[0] for item in items]
    row_gap = max((max(y_values) - min(y_values)) / max(len(y_values), 1) * 1.5, 20.0)

    rows: list[list[tuple[float, float, str]]] = []
    current: list[tuple[float, float, str]] = [items[0]]
    for item in items[1:]:
        if item[0] - current[-1][0] <= row_gap:
            current.append(item)
        else:
            rows.append(current)
            current = [item]
    rows.append(current)

    # 各行内を x 座標でソート（左→右）
    return [[text for _, _, text in sorted(row, key=lambda r: r[1])] for row in rows]
